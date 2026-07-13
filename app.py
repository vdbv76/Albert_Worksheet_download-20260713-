"""
Albert Invent - Worksheet Duplicate (v3)
========================================

Faithful recreation of the Albert Worksheet in Streamlit via the official
`albert` SDK, for ANY project (MO13137 was used as the reference example).

v3 changes (driven by the "Worksheet structure" README):
  1. GLOBAL FILTER PANEL - re-implements Albert's 7 client-side UI filters
     (README §9: they are NOT persisted server-side, the API always returns
     the full set, so they must be rebuilt client-side):
        - Formula / Product ID (text)
        - Contains ingredient ("Inventory" filter - Product Design INV rows)
        - Locked / Unlocked
        - Predecessor
        - Tags
        - Data Templates
        - Created By
     One filter state -> ONE set of visible experiment columns applied to ALL
     four section tables (Product / Results / Apps / Process) AND downloads.
  2. Tags / Predecessor / Created-By come from InventoryItem entities
     (inventory.get_by_ids - batched 250/call), NOT from the grid TAG row,
     which is unreliable on large sheets (README §16.2).
  3. Data-Template membership comes from tasks.search (one call - each
     TaskSearchItem already carries its dataTemplate + inventory lists).
  4. Results are still loaded lazily per selected Property Task via
     property_data.get_all_task_properties (README §12: fastest bulk path,
     one logical call per task).
  5. interval_combination tokens (ROW4, ROW4XROW2) are resolved to human-
     readable parameter setpoints via workflows.get_by_ids ->
     Workflow.interval_combinations (README §11) - the ROW tokens are
     positional indices inside the workflow, NOT worksheet row ids.
  6. Focus view (hide rows empty across the visible columns), README §9.

Requirements:
    pip install streamlit albert pandas truststore openpyxl

Run:
    streamlit run app.py
"""

from __future__ import annotations

# --- Corporate SSL fix: must run before any HTTPS request -------------------
import truststore

truststore.inject_into_ssl()

import io
import os
import re
from typing import Any

import pandas as pd
import streamlit as st

from albert import Albert

st.set_page_config(page_title="Albert - Worksheet Duplicate", layout="wide")
st.title("🧪 Albert Worksheet - Live Duplicate")

SECTION_ORDER = [
    ("product_design", "Product Design"),
    ("result_design", "Results"),
    ("app_design", "Apps"),
    ("process_design", "Process Design"),
]

# Sentinel option in every filter: matches items that have NO value for that
# attribute (no tag, no predecessor, no group at that level, ...). Without it,
# selecting every real option still silently drops the blanks.
NONE_LABEL = "(None)"

ROW_TYPE_LABELS = {
    "INV": "Inventory",
    "BLK": "Blank",
    "TAS": "Task",
    "PRG": "Param. Group",
    "PRM": "Parameter",
    "TOT": "Total",
    "TAG": "Tags",
    "APP": "App link",
    "PRC": "Pricing",
    "PDC": "Predecessor",
    "BAT": "Batches",
    "LKP": "Lookup",
    "RSL": "Substance data",
    "DEF": "Default",
    "Formula": "Formula",
    "DAT": "Data Template",
    "DAC": "Data Column",
}


# ===========================================================================
# Helpers
# ===========================================================================
def _cell_text(cell: Any) -> str:
    if cell is None:
        return ""
    v = getattr(cell, "value", None)
    if v in (None, ""):
        return ""
    if isinstance(v, dict):
        for k in ("value", "name", "text"):
            if v.get(k) not in (None, ""):
                return str(v[k])
        return ""
    if isinstance(v, list):
        return ", ".join(str(x) for x in v if x not in (None, ""))
    return str(v)


def _friendly_type(raw: str) -> str:
    raw = str(raw or "")
    return ROW_TYPE_LABELS.get(raw.split(".")[-1], raw.split(".")[-1])


def _strip_inv(row_link_id: str | None) -> str:
    """Grid row ids are prefixed with 'INV'; strip to reveal the entity id
    (e.g. INVTAS123 -> TAS123)."""
    if not row_link_id:
        return ""
    return row_link_id[3:] if row_link_id.startswith("INV") else row_link_id


def _to_number(x: Any) -> tuple[Any, bool]:
    """(float, True) when x parses as a number (comma OR dot decimal), else (x, False)."""
    try:
        return float(str(x).strip().replace(",", ".")), True
    except (ValueError, AttributeError):
        return x, False


def _cmp_pass(value: Any, mode: str, a: Any, b: Any = None) -> bool:
    """Generic column filter (Product-Design Name, Results Data Column).
    Compares numerically when both sides parse as numbers, lexicographically
    otherwise. mode is one of All / Contains / > / < / Between."""
    if mode in (None, "", "All"):
        return True
    if a in (None, ""):
        return True
    v = str(value)
    if mode == "Contains":
        return str(a).lower() in v.lower()
    va, vnum = _to_number(v)
    aa, anum = _to_number(a)
    if mode == ">":
        return (va > aa) if (vnum and anum) else (v > str(a))
    if mode == "<":
        return (va < aa) if (vnum and anum) else (v < str(a))
    if mode == "Between":
        if b in (None, ""):
            return True
        bb, bnum = _to_number(b)
        if vnum and anum and bnum:
            lo, hi = (aa, bb) if aa <= bb else (bb, aa)
            return lo <= va <= hi
        lo, hi = (str(a), str(b)) if str(a) <= str(b) else (str(b), str(a))
        return lo <= v <= hi
    return True


def _round_sig_decimals(tok: str, decimals: int) -> str:
    """Round ONE numeric token keeping `decimals` significant fractional digits,
    counted from the first non-zero decimal place:
        0.0012345 -> 0.0012   (decimals=2)
        1.0123    -> 1.012
        1.123     -> 1.12
    Integers and non-numeric tokens are returned unchanged."""
    t = tok.strip()
    if t == "":
        return tok
    comma = "," in t and "." not in t
    norm = t.replace(",", ".") if comma else t
    try:
        val = float(norm)
    except ValueError:
        return tok
    if "." not in norm:  # an integer -> leave it alone
        return tok
    frac = norm.split(".", 1)[1]
    lz = 0
    for ch in frac:
        if ch == "0":
            lz += 1
        else:
            break
    out = f"{val:.{lz + decimals}f}"
    return out.replace(".", ",") if comma else out


def _apply_decimals_text(text: Any, decimals: int | None) -> Any:
    """Apply `_round_sig_decimals` to every numeric token in a cell, preserving the
    ' | ' separators used to list repeated measurements."""
    if decimals is None:
        return text
    s = str(text)
    if s == "":
        return s
    if "|" not in s:
        return _round_sig_decimals(s, decimals)
    return " | ".join(_round_sig_decimals(p.strip(), decimals) for p in s.split("|"))


# ===========================================================================
# Sidebar: authentication
# ===========================================================================
with st.sidebar:
    st.header("🔐 Connect to Albert")
    base_url = st.text_input("Base URL", value="https://app.albertinvent.com")
    method = st.radio("Authentication", ["SSO login", "Client Credentials", "Static token"])
    if method == "SSO login":
        email = st.text_input("Albert account email")
        if st.button("Connect via SSO", type="primary", use_container_width=True):
            try:
                with st.spinner("Complete the login in the opened browser tab..."):
                    st.session_state["client"] = Albert.from_sso(base_url=base_url, email=email)
                st.success("Connected ✅")
            except Exception as e:  # noqa: BLE001
                st.error(f"Could not connect: {e}")
    elif method == "Client Credentials":
        cid = st.text_input("Client ID")
        sec = st.text_input("Client Secret", type="password")
        if st.button("Connect", type="primary", use_container_width=True):
            try:
                st.session_state["client"] = Albert.from_client_credentials(
                    base_url=base_url, client_id=cid, client_secret=sec
                )
                st.success("Connected ✅")
            except Exception as e:  # noqa: BLE001
                st.error(f"Could not connect: {e}")
    else:
        tok = st.text_input("Token (JWT)", type="password")
        if st.button("Connect", type="primary", use_container_width=True):
            try:
                st.session_state["client"] = Albert.from_token(base_url=base_url, token=tok)
                st.success("Connected ✅")
            except Exception as e:  # noqa: BLE001
                st.error(f"Could not connect: {e}")

client: Albert | None = st.session_state.get("client")
if client is None:
    st.info("👈 Connect to Albert from the sidebar to get started.")
    st.stop()


# ===========================================================================
# 1) Project & sheet selection
# ===========================================================================
st.header("1️⃣ Project & sheet")

c1, c2 = st.columns([3, 1])
with c1:
    q = st.text_input("Search project (e.g. MO13137)")
with c2:
    only_mine = st.checkbox("Only my projects")

if st.button("🔎 Search"):
    try:
        with st.spinner("Searching..."):
            st.session_state["projects"] = list(
                client.projects.search(
                    text=q or None,
                    my_project=True if only_mine else None,
                    max_items=50,
                )
            )
    except Exception as e:  # noqa: BLE001
        st.error(f"Search failed: {e}")

projects = st.session_state.get("projects", [])
if not projects:
    st.stop()

proj_labels = {f"{p.description}  [{p.id}]": p.id for p in projects}
sel_proj = st.selectbox("Project", list(proj_labels.keys()))
project_id = proj_labels[sel_proj]


@st.cache_resource(show_spinner="Loading worksheet from Albert...")
def get_worksheet(_client: Albert, pid: str):
    return _client.worksheets.get_by_project_id(project_id=pid)


try:
    worksheet = get_worksheet(client, project_id)
except Exception as e:  # noqa: BLE001
    st.error(f"Could not load the worksheet: {e}")
    st.stop()

sheets = worksheet.sheets or []
if not sheets:
    st.warning("This project's worksheet has no sheets.")
    st.stop()

sheet_names = {getattr(s, "name", "") or "(unnamed)": s for s in sheets}
sel_sheet_name = st.selectbox("Sheet", list(sheet_names.keys()))
sheet = sheet_names[sel_sheet_name]


# ===========================================================================
# 2) Extract sheet structure (grid values + column metadata incl. locked)
# ===========================================================================
@st.cache_data(show_spinner="Reading sheet grid...")
def extract_sheet(_sheet, sheet_key: str) -> dict:
    def _row_hierarchy(design) -> dict:
        """rowId -> ordered ancestor names [Group, Subgroup 1, ..., Subgroup n].

        SOURCE OF TRUTH: GET /api/v3/worksheet/design/{id}/rows/sequence.
        Row grouping in Albert is an explicit server-side parent->child
        relationship (Design.group_rows() PUTs an explicit ChildRows list), so
        the tree is stored, not inferred. The SDK's own get_groups() reads only
        ONE level of that response and throws away nested subgroups - so we
        call the endpoint directly and walk it recursively to arbitrary depth.

        The grid/sheet_inspect endpoint is a FLAT array with no parent, child,
        depth or indent field. Depth CANNOT be recovered from it: a BLK row
        followed by another BLK row is ambiguous between "child" and "sibling",
        and that ambiguity is not resolvable by any rule. We therefore do NOT
        guess. If the sequence endpoint is unavailable, hierarchy is reported
        as UNAVAILABLE rather than fabricated.

        Returns {"paths": {rid: [ancestor_rid, ...]}, "node_names": {rid: name},
                 "source": str, "raw": <json|None>, "error": str|None, "keys": {...}}

        NOTE: the sequence tree identifies nodes by rowId and does not reliably
        carry a display name. Ancestor NAMES are therefore resolved afterwards
        from the grid's rowId -> name map (BLK group headers appear in the grid
        with their labels, e.g. ROW338 -> "Raw Materials"). Any name the tree
        does supply is kept in node_names and wins over the grid.
        """
        out = {
            "paths": {},
            "node_names": {},
            "source": "unavailable",
            "raw": None,
            "error": None,
            "keys": {},
        }

        seq = None
        try:
            resp = design.session.get(f"/api/v3/worksheet/design/{design.id}/rows/sequence")
            seq = resp.json()
            out["raw"] = seq
        except Exception as e:  # noqa: BLE001
            out["error"] = f"{type(e).__name__}: {e}"

        ID_KEYS = ("rowId", "row_id", "id")
        NAME_KEYS = ("name", "lableName", "labelName", "rowName", "label")
        KID_KEYS = ("children", "childRows", "ChildRows", "Children", "Rows", "rows")

        def first(node: dict, keys) -> Any:
            for k in keys:
                v = node.get(k)
                if v not in (None, "", []):
                    out["keys"][k] = out["keys"].get(k, 0) + 1
                    return v
            return None

        def kids(node: dict) -> list:
            for k in KID_KEYS:
                v = node.get(k)
                if isinstance(v, list) and v:
                    out["keys"][k] = out["keys"].get(k, 0) + 1
                    return v
            return []

        paths: dict[str, list[str]] = {}
        nested = False

        def walk(nodes, ancestor_ids: list[str]):
            nonlocal nested
            for n in nodes:
                if not isinstance(n, dict):
                    continue
                rid = first(n, ID_KEYS)
                if not rid:
                    continue
                rid = str(rid)
                nm = first(n, NAME_KEYS)
                if nm:
                    out["node_names"][rid] = str(nm)
                paths[rid] = ancestor_ids[:]
                ch = kids(n)
                if ch:
                    nested = True
                    walk(ch, ancestor_ids + [rid])

        if isinstance(seq, list) and seq:
            walk(seq, [])
        elif isinstance(seq, dict):
            for k in ("Items", "items", "Rows", "rows", "data", "Data"):
                if isinstance(seq.get(k), list):
                    walk(seq[k], [])
                    break

        if paths and nested:
            out["paths"] = paths
            out["source"] = "rows/sequence endpoint (explicit parent->child tree)"
            return out

        # The endpoint returned nothing usable. Last resort: the SDK's
        # one-level get_groups() (same endpoint, shallow parse) - gives a
        # single Group level only, never subgroups. Still real data, not a guess.
        try:
            groups = design.get_groups()
        except Exception:  # noqa: BLE001
            groups = []
        if groups:
            for g in groups:
                if g.name:
                    out["node_names"][g.row_id] = g.name
                for ch in g.child_row_ids:
                    paths[ch] = [g.row_id]
            out["paths"] = paths
            out["source"] = "get_groups() - ONE LEVEL ONLY (no subgroups available)"
            return out

        out["source"] = "unavailable"
        return out

    inv_to_form_name = {
        f.id: (f.name or "")
        for f in (getattr(_sheet, "formulations", None) or [])
        if getattr(f, "id", None)
    }

    columns = []
    for c in _sheet.columns:
        name = getattr(c, "name", None) or ""
        columns.append(
            {
                "column_id": getattr(c, "column_id", None),
                "name": name,
                "type": str(getattr(c, "type", "") or ""),
                "inventory_id": getattr(c, "inventory_id", None),
                "hidden": bool(getattr(c, "hidden", False)),
                "locked": bool(getattr(c, "locked", False)),
                "pinned": getattr(c, "pinned", None),
                "formulation_name": inv_to_form_name.get(getattr(c, "inventory_id", None), ""),
                # The sheet's built-in label column duplicates the row names -
                # exclude it from the data columns.
                "is_label_col": name.strip().lower() == "name",
            }
        )

    sections = []
    for attr, label in SECTION_ORDER:
        design = getattr(_sheet, attr, None)
        if design is None:
            continue
        try:
            grid = design.grid
        except Exception:  # noqa: BLE001
            continue
        if grid is None or grid.empty:
            continue

        # FIX: each Design parses its OWN columns from its OWN grid response
        # (Process Design even uses a different endpoint), so column_ids are NOT
        # comparable across designs - Sheet.columns is product_design.columns only.
        # Key every cell by inventory_id, which IS stable across designs. This is
        # what makes one filter state apply identically to all four sections.
        try:
            design_col_inv = {
                getattr(c, "column_id", None): getattr(c, "inventory_id", None)
                for c in design.columns
            }
        except Exception:  # noqa: BLE001
            design_col_inv = {}

        # Full nested hierarchy (Group / Subgroup 1 / ... / Subgroup n)
        hier = _row_hierarchy(design)
        row_paths = hier["paths"]

        rows = []
        for _, row_series in grid.iterrows():
            first_cell = next((c for _, c in row_series.items() if c is not None), None)
            if first_cell is None:
                continue
            rid = getattr(first_cell, "row_id", None)
            label_name = (
                getattr(first_cell, "row_label_name", None)
                or getattr(first_cell, "name", None)
                or ""
            )
            rtype_raw = str(getattr(first_cell, "row_type", "") or "")
            link_id = _strip_inv(getattr(first_cell, "inventory_id", None))
            values: dict[str, str] = {}  # inventory_id -> cell text
            for _, cell in row_series.items():
                if cell is None or (isinstance(cell, float) and pd.isna(cell)):
                    continue
                cid = getattr(cell, "column_id", None)
                # NB: cell.inventory_id is the ROW's inventory item (the raw
                # material), never the column's - only the design column map is valid.
                inv = design_col_inv.get(cid)
                if inv:
                    values[inv] = _cell_text(cell)
            path_ids = list(row_paths.get(rid, []))  # [ancestor rowIds, outer->inner]
            rows.append(
                {
                    "row_id": rid,
                    "name": label_name,
                    "type_raw": rtype_raw,
                    "type": _friendly_type(rtype_raw),
                    "path_ids": path_ids,
                    "path": [],  # filled in below, once every row name is known
                    "depth": len(path_ids) + 1,
                    "link_id": link_id,  # e.g. TAS123 for Property Block rows
                    "values": values,
                }
            )

        # --- resolve ancestor rowIds -> display names -------------------------
        # The sequence tree is keyed by rowId and carries no reliable label, so
        # the group/subgroup names come from the grid itself: a group header IS
        # a BLK row, and that row's name is its label (ROW338 -> "Raw Materials").
        name_by_rid = {r["row_id"]: r["name"] for r in rows if r["row_id"]}
        node_names = hier["node_names"]  # names the tree did supply, if any
        unresolved: set[str] = set()
        for r in rows:
            names = []
            for aid in r["path_ids"]:
                nm = node_names.get(aid) or name_by_rid.get(aid) or ""
                if not nm:
                    unresolved.add(aid)
                    nm = str(aid)  # last resort: show the id rather than blank
                names.append(nm)
            r["path"] = names

        if rows:
            sections.append(
                {
                    "attr": attr,
                    "label": label,
                    "rows": rows,
                    "hierarchy_source": hier["source"],
                    "hierarchy_error": hier["error"],
                    "hierarchy_keys": hier["keys"],
                    "hierarchy_raw": hier["raw"],
                    "hierarchy_unresolved": sorted(unresolved),
                    "max_depth": max((len(r["path"]) for r in rows), default=0),
                }
            )

    return {"columns": columns, "sections": sections}


data = extract_sheet(sheet, f"{project_id}::{sel_sheet_name}")
columns, sections = data["columns"], data["sections"]

if not sections:
    st.warning("No grid data found in this sheet.")
    st.stop()

section_by_attr = {s["attr"]: s for s in sections}


# ===========================================================================
# 3) Enrichment - the data behind the filters
#    (README §9/§14/§16: tags, predecessor and creator come from Inventory
#    entities, NOT the grid; Data-Template membership comes from task search)
# ===========================================================================
def _tag_name(t: Any) -> str:
    """A Tag's NAME lives in `Tag.tag` (alias 'name'/'tagName'), NOT `Tag.name`.
    Reading `.name` silently yields None for every tag - which is why the Tags
    filter came up empty. Falls back through the other shapes defensively."""
    if isinstance(t, str):
        return t
    for attr in ("tag", "name", "tag_name"):
        v = getattr(t, attr, None)
        if v:
            return str(v)
    if isinstance(t, dict):
        for k in ("tag", "name", "tagName"):
            if t.get(k):
                return str(t[k])
    return ""


def _tag_id(t: Any) -> str:
    for attr in ("id", "albert_id", "tag_id"):
        v = getattr(t, attr, None)
        if v:
            return str(v)
    if isinstance(t, dict):
        return str(t.get("id") or t.get("albertId") or "")
    return ""


@st.cache_data(show_spinner="Loading formulation metadata (tags, creators)...")
def load_inventory_meta(_client: Albert, inv_ids: tuple[str, ...]) -> dict[str, dict]:
    """inventory_id -> {name, alias, description, tags, created_by}.

    TAGS: `get_by_ids` returns tag links that may carry only the TAG id, so any
    id whose name is missing is resolved in one batched `tags.get_by_ids` call.
    PREDECESSOR is NOT on the InventoryItem at all (no top-level field, not in
    Metadata, no facet) - it comes from the worksheet's Apps PDC row instead;
    see `predecessor_by_inv()`.
    """
    out: dict[str, dict] = {}
    ids = [i for i in inv_ids if i]
    if not ids:
        return out
    try:
        items = _client.inventory.get_by_ids(ids=list(ids))
    except Exception as e:  # noqa: BLE001
        st.warning(f"Could not load inventory metadata (filters degraded): {e}")
        return out

    unresolved_tag_ids: set[str] = set()
    for it in items:
        tag_names, tag_ids = [], []
        for t in getattr(it, "tags", None) or []:
            nm, tid = _tag_name(t), _tag_id(t)
            if nm:
                tag_names.append(nm)
            elif tid:
                tag_ids.append(tid)
                unresolved_tag_ids.add(tid)
        created = getattr(it, "created", None)
        out[it.id] = {
            "name": getattr(it, "name", "") or "",
            "alias": getattr(it, "alias", "") or "",
            "description": getattr(it, "description", "") or "",
            "tags": tag_names,
            "_tag_ids": tag_ids,  # names still to resolve
            "created_by": getattr(created, "by_name", None) or getattr(created, "by", "") or "",
            "created_at": str(getattr(created, "at", "") or ""),
        }

    # Resolve any id-only tags -> names (one batched call)
    if unresolved_tag_ids:
        id_to_name: dict[str, str] = {}
        try:
            for t in _client.tags.get_by_ids(ids=sorted(unresolved_tag_ids)):
                id_to_name[_tag_id(t)] = _tag_name(t)
        except Exception:  # noqa: BLE001
            pass
        for m in out.values():
            for tid in m["_tag_ids"]:
                nm = id_to_name.get(tid)
                m["tags"].append(nm if nm else tid)

    for m in out.values():
        m["tags"] = sorted(set(t for t in m["tags"] if t))
        m.pop("_tag_ids", None)
    return out


@st.cache_data(show_spinner="Loading filter facets...")
def load_facets(_client: Albert, pid: str) -> dict[str, list[tuple[str, int]]]:
    """Albert's own filter-dropdown source: inventory facets, project-scoped.
    parameter -> [(value, count), ...]. Used to seed the Tags / Created By
    dropdowns so they show every value that exists, with counts."""
    out: dict[str, list[tuple[str, int]]] = {}
    try:
        for f in _client.inventory.get_all_facets(project_id=pid):
            out[f.parameter] = [(v.name, v.count) for v in (f.value or [])]
    except Exception:  # noqa: BLE001
        pass
    return out


@st.cache_data(show_spinner="Listing Property Tasks...")
def get_property_tasks(_client: Albert, pid: str) -> list[dict]:
    """One search call. Each TaskSearchItem already carries its data-template
    names and inventory ids - enough for the Data-Templates filter without
    hydrating any task."""
    out = []
    try:
        for t in _client.tasks.search(project_id=pid, category="Property", max_items=500):
            out.append(
                {
                    "id": t.id,
                    "name": getattr(t, "name", "") or "",
                    "state": getattr(t, "state", "") or "",
                    "data_templates": [
                        dt.name for dt in (getattr(t, "data_template", None) or []) if dt.name
                    ],
                    "data_template_ids": [
                        dt.id for dt in (getattr(t, "data_template", None) or []) if dt.id
                    ],
                    "inventory_ids": [
                        inv.id for inv in (getattr(t, "inventory", None) or []) if inv.id
                    ],
                }
            )
    except Exception as e:  # noqa: BLE001
        st.warning(f"Task search failed: {e}")
    return out


@st.cache_data(show_spinner="Resolving data template names...")
def load_data_template_names(_client: Albert, dt_ids: tuple[str, ...]) -> dict[str, str]:
    """DataTemplate id -> canonical short name. The name string carried on the
    task-search response is not necessarily the one the Worksheet dropdown shows
    (DataTemplate has name / fullName / originalName, e.g. DAT235 is 'Cobb Value'
    but fullName is 'DIN EN 20535: Cobb Value'). Resolve from the entity itself."""
    out: dict[str, str] = {}
    ids = [i for i in dt_ids if i]
    if not ids:
        return out
    try:
        for dt in _client.data_templates.get_by_ids(ids=list(ids)):
            full = getattr(dt, "full_name", None)
            nm = getattr(dt, "name", "") or ""
            out[dt.id] = f"{nm} ({full})" if full and full != nm else nm
    except Exception:  # noqa: BLE001
        pass
    return out


exp_inventory_ids = tuple(
    c["inventory_id"] for c in columns if c["inventory_id"] and not c["is_label_col"]
)
inv_meta = load_inventory_meta(client, exp_inventory_ids)
facets = load_facets(client, project_id)
property_tasks = get_property_tasks(client, project_id)

# --- PREDECESSOR: only lives in the worksheet's Apps design PDC row ----------
# It is NOT a field on InventoryItem (not top-level, not in Metadata, no facet).
# We read the Apps grid via the per-design /grid endpoint, which is NOT subject
# to the 20k-item truncation that hits sheets.get_cell_values on large sheets.
def _apps_row_values(row_type: str) -> dict[str, str]:
    sec = section_by_attr.get("app_design")
    if not sec:
        return {}
    for r in sec["rows"]:
        if r["type_raw"].split(".")[-1].upper() == row_type:
            return {inv: v for inv, v in r["values"].items() if v}
    return {}


predecessor_by_inv = _apps_row_values("PDC")
for _inv, _m in inv_meta.items():
    _m["predecessor"] = predecessor_by_inv.get(_inv, "")

# inventory_id -> set of data-template names (via the Property Tasks it appears in)
dt_ids_all = tuple({i for t in property_tasks for i in t["data_template_ids"]})
dt_name_of = load_data_template_names(client, dt_ids_all)
dts_of_inv: dict[str, set[str]] = {}
for t in property_tasks:
    names = {
        dt_name_of.get(i) or nm
        for i, nm in zip(t["data_template_ids"], t["data_templates"])
    }
    for inv in t["inventory_ids"]:
        dts_of_inv.setdefault(inv, set()).update(n for n in names if n)
all_data_templates = sorted({dt for s in dts_of_inv.values() for dt in s})


def column_header(c: dict) -> tuple[str, str]:
    """(top, bottom) header: top = short code (e.g. MO13137-053), bottom = name."""
    if not c["inventory_id"]:
        return (c["name"] or c["column_id"] or "", "")
    meta = inv_meta.get(c["inventory_id"], {})
    long_name = c["name"] or meta.get("name") or c["formulation_name"]
    candidates = [meta.get("alias", ""), meta.get("name", ""), c["formulation_name"]]
    code = ""
    for cand in candidates:
        if cand and cand != long_name and len(cand) <= 40:
            code = cand
            break
    if not code:
        code = _strip_inv(c["inventory_id"])
    return (code, long_name)


# ===========================================================================
# 4) GLOBAL FILTERS - one state, applied to every section table + downloads
#    Re-implementation of Albert's 7 client-side UI filters (README §9).
# ===========================================================================
st.header("2️⃣ Filters (apply to all sections)")

exp_cols_all = [c for c in columns if c["inventory_id"] and not c["is_label_col"]]

# Ingredient candidates = Product Design INV rows ("Contains inventory" filter)
product_section = section_by_attr.get("product_design")
inv_rows_product = (
    [r for r in product_section["rows"] if r["type_raw"].split(".")[-1] == "INV"]
    if product_section
    else []
)
ingredient_options = sorted({r["name"] for r in inv_rows_product if r["name"]})


def _with_none(options: list[str], has_blank: bool) -> list[str]:
    """Prepend the (None) sentinel when some column has no value for this attribute."""
    return ([NONE_LABEL] if has_blank else []) + options


def _options_from(attr: str, facet_param: str | None = None) -> list[str]:
    """Filter options = every value present on THIS sheet's columns, unioned with
    Albert's own facet list for the project (so nothing that exists is missing).
    Sorted by how many of the visible experiment columns carry it."""
    counts: dict[str, int] = {}
    for c in exp_cols_all:
        m = inv_meta.get(c["inventory_id"], {})
        vals = m.get(attr) or []
        if isinstance(vals, str):
            vals = [vals] if vals else []
        for v in vals:
            counts[v] = counts.get(v, 0) + 1
    if facet_param:
        for name, _cnt in facets.get(facet_param, []):
            counts.setdefault(name, 0)
    return sorted(counts, key=lambda v: (-counts[v], v.lower()))


all_tags = _with_none(
    _options_from("tags", "tags"),
    any(not m.get("tags") for m in inv_meta.values()),
)
all_creators = _with_none(
    _options_from("created_by", "createdBy"),
    any(not m.get("created_by") for m in inv_meta.values()),
)
all_predecessors = _with_none(
    _options_from("predecessor"),
    any(not m.get("predecessor") for m in inv_meta.values()),
)

f1, f2, f3 = st.columns(3)
with f1:
    # FIX #2: Albert's Formula/Product ID filter is a searchable dropdown, not a
    # free-text box. Streamlit's multiselect does substring type-ahead over the
    # option labels, so typing "MO13137-09" or "85p (PA" narrows the list live.
    exp_options = {
        f"{code}  ·  {desc}" if desc else code: c["inventory_id"]
        for c, (code, desc) in zip(
            exp_cols_all, [column_header(c) for c in exp_cols_all]
        )
    }
    flt_exp_labels = st.multiselect(
        "Formula / Product ID",
        list(exp_options.keys()),
        help="Type any part of the ID or the name (e.g. 'MO13137-09' or '85p (PA') "
        "to narrow the list, then pick the experiments you want.",
    )
    flt_exp_invs = {exp_options[l] for l in flt_exp_labels}
    flt_ingredients = st.multiselect(
        "Contains ingredient (Product Design)",
        ingredient_options,
        help="Show only formulations that have a value in the selected ingredient row(s).",
    )
with f2:
    flt_tags = st.multiselect(
        "Tags", all_tags, help=f"'{NONE_LABEL}' = formulations with no tag."
    )
    flt_creators = st.multiselect(
        "Created by", all_creators, help=f"'{NONE_LABEL}' = creator unknown."
    )
with f3:
    flt_preds = st.multiselect(
        "Predecessor",
        all_predecessors,
        help="Formulations derived from the selected predecessor formula(s). "
        f"'{NONE_LABEL}' = formulations with no predecessor.",
    )
    # FIX #3: Data Templates is a RESULTS-ROW filter, not a column filter. It is
    # only meaningful once Property Tasks have been loaded, and selecting one must
    # never hide experiment columns from the other three sections.
    loaded_dts = sorted(
        {
            r["Data Template"]
            for recs in st.session_state.get(f"results_store::v3::{project_id}", {}).values()
            for r in recs
            if "__error__" not in r and r.get("Data Template")
        }
    )
    flt_result_dts = st.multiselect(
        "Data Templates (Results only)",
        loaded_dts,
        disabled=not loaded_dts,
        help="Filters the rows of the Results tables. Enabled once you load at least "
        "one Property Task below. It does not hide experiment columns."
        if loaded_dts
        else "Load a Property Task in the Results section first.",
    )

f4, f5, f6, f7 = st.columns(4)
with f4:
    match_all = st.checkbox(
        "Match ALL conditions within a filter",
        value=False,
        help="e.g. must contain every selected ingredient / carry every selected tag.",
    )
with f5:
    flt_lock = st.radio("Lock state", ["All", "Locked", "Unlocked"], horizontal=True)
with f6:
    show_hidden = st.checkbox("Show hidden columns", value=False)
with f7:
    focus_view = st.checkbox(
        "Hide empty rows (Focus view)",
        value=False,
        help="Hide rows that have no value in any of the visible experiment columns.",
    )


def _ingredient_hit(col: dict, wanted: list[str]) -> bool:
    inv = col["inventory_id"]
    hits = [
        any(r["name"] == w and r["values"].get(inv, "") != "" for r in inv_rows_product)
        for w in wanted
    ]
    return all(hits) if match_all else any(hits)


def _set_filter_hit(have: set[str], wanted: list[str]) -> bool:
    """(None) matches items with an empty set; real selections use the ANY/ALL
    toggle. The two are OR'd, so selecting every option shows every column."""
    real = [w for w in wanted if w != NONE_LABEL]
    if not have:
        return NONE_LABEL in wanted
    if not real:
        return False  # only (None) was selected, and this item has values
    return have.issuperset(real) if match_all else bool(have & set(real))


def _scalar_filter_hit(have: str, wanted: list[str]) -> bool:
    if not have:
        return NONE_LABEL in wanted
    return have in wanted


def column_passes(c: dict) -> bool:
    if not show_hidden and c["hidden"]:
        return False
    if flt_lock == "Locked" and not c["locked"]:
        return False
    if flt_lock == "Unlocked" and c["locked"]:
        return False

    meta = inv_meta.get(c["inventory_id"], {})

    if flt_exp_invs and c["inventory_id"] not in flt_exp_invs:
        return False

    if flt_ingredients and not _ingredient_hit(c, flt_ingredients):
        return False

    if flt_tags and not _set_filter_hit(set(meta.get("tags", [])), flt_tags):
        return False

    if flt_creators and not _scalar_filter_hit(meta.get("created_by", ""), flt_creators):
        return False

    if flt_preds and not _scalar_filter_hit(meta.get("predecessor", ""), flt_preds):
        return False

    return True


visible_cols = [c for c in exp_cols_all if column_passes(c)]

n_hidden = sum(1 for c in exp_cols_all if c["hidden"])
st.caption(
    f"**{len(visible_cols)} / {len(exp_cols_all)} experiment columns** pass the filters"
    + (f" ({n_hidden} hidden in Albert)" if not show_hidden and n_hidden else "")
)
if not visible_cols:
    st.warning("No experiment column passes the current filters.")
    st.stop()

col_tuples = []
_seen_codes: dict[str, int] = {}
for _c in visible_cols:
    _code, _desc = column_header(_c)
    _code = _code or _strip_inv(_c["inventory_id"])
    if _code in _seen_codes:  # codes must be unique - they are the column labels
        _seen_codes[_code] += 1
        _code = f"{_code} ({_seen_codes[_code]})"
    else:
        _seen_codes[_code] = 1
    col_tuples.append((_code, _desc))

colid_to_tuple = dict(zip([c["column_id"] for c in visible_cols], col_tuples))
invid_to_tuple = {
    c["inventory_id"]: t for c, t in zip(visible_cols, col_tuples) if c["inventory_id"]
}


# ===========================================================================
# 5) Row display options + section tables
# ===========================================================================
st.header("3️⃣ Worksheet")

# Group / Subgroup columns are always built now; they can be hidden per-table via
# each table's "Hide columns" dropdown instead of a single global toggle.
show_hier_cols = True

o1, o2, o3, o4 = st.columns(4)
with o1:
    dec_choice = st.selectbox(
        "Decimals",
        ["All"] + list(range(0, 7)),
        index=0,
        help="Round numbers to this many significant decimals, counted from the first "
        "non-zero decimal place. e.g. Decimals=2 turns 0.0012345 -> 0.0012, "
        "1.0123 -> 1.012 and 1.123 -> 1.12. 'All' leaves numbers untouched.",
    )
    DECIMALS = None if dec_choice == "All" else int(dec_choice)
with o2:
    show_type_col = st.checkbox("Show 'Row type' column", value=False)
with o3:
    hide_blk = st.checkbox("Hide Blank (BLK) rows", value=False)
with o4:
    indent_names = st.checkbox(
        "Indent row names by depth", value=False, help="Mimics the Albert UI tree."
    )
    show_desc_row = st.checkbox(
        "Show experiment description row",
        value=True,
        help="Adds the full formulation name as the first row. The name is also "
        "available as a tooltip on each column header.",
    )


def hier_cols_for(section: dict) -> list[str]:
    """['Group', 'Subgroup 1', ..., 'Subgroup n'] sized to this section's tree."""
    if not show_hier_cols or not section["max_depth"]:
        return []
    d = section["max_depth"]
    return ["Group"] + [f"Subgroup {i}" for i in range(1, d)]


def key_cols_for(section: dict) -> list[str]:
    return (
        ["Name"]
        + hier_cols_for(section)
        + (["Row type"] if show_type_col else [])
    )


def _row_in_filter(r: dict, row_filter: dict[int, list[str]]) -> bool:
    """row_filter: {ancestor_level -> allowed names}. Empty selection at a level
    = no filtering at that level. NONE_LABEL matches rows that have no ancestor
    at that level (i.e. the cell is blank), so 'select everything' really does
    mean everything. A group header row is kept when its own name is selected."""
    for level, wanted in row_filter.items():
        if not wanted:
            continue
        at_level = r["path"][level] if len(r["path"]) > level else ""
        # a header row sits AT this level: its own name is the value shown below it
        is_the_header = len(r["path"]) == level and r["name"] in wanted
        if at_level == "":
            if NONE_LABEL in wanted or is_the_header:
                continue
            return False
        if at_level not in wanted:
            return False
    return True


def _cmp_filter_widget(label: str, key: str, help_txt: str = "") -> tuple[str, str, str]:
    """Render a compact 'column filter' (mode + up to two values) and return
    (mode, a, b). Used by the Product-Design Name and Results Data Column filters."""
    cols = st.columns([1.4, 1, 1])
    with cols[0]:
        mode = st.selectbox(
            label,
            ["All", "Contains", ">", "<", "Between"],
            key=f"{key}::mode",
            help=help_txt or None,
        )
    a = b = ""
    if mode != "All":
        with cols[1]:
            a = st.text_input("From" if mode == "Between" else "Value", key=f"{key}::a")
    if mode == "Between":
        with cols[2]:
            b = st.text_input("To", key=f"{key}::b")
    return mode, a, b


def rows_dataframe(
    section: dict,
    row_filter: dict | None = None,
    with_ids: bool = False,
    name_cmp: tuple[str, str, str] | None = None,
):
    hcols = hier_cols_for(section)
    kcols = key_cols_for(section)
    recs, rids = [], []
    for r in section["rows"]:
        if hide_blk and r["type_raw"].split(".")[-1] == "BLK":
            continue
        if row_filter and not _row_in_filter(r, row_filter):
            continue
        if name_cmp and not _cmp_pass(r["name"], *name_cmp):
            continue
        vals = {c["inventory_id"]: r["values"].get(c["inventory_id"], "") for c in visible_cols}
        if focus_view and not any(v != "" for v in vals.values()):
            continue

        name = r["name"]
        if indent_names and r["path"]:
            name = ("\u00a0" * 4 * len(r["path"])) + name
        rec = {"Name": name}
        for i, hc in enumerate(hcols):
            rec[hc] = r["path"][i] if len(r["path"]) > i else ""
        if show_type_col:
            rec["Row type"] = r["type"]
        for c, t in zip(visible_cols, col_tuples):
            rec[t] = vals[c["inventory_id"]]
        recs.append(rec)
        rids.append(str(r["row_id"]))
    df = pd.DataFrame(recs).reindex(columns=kcols + col_tuples).fillna("")
    return (df, rids) if with_ids else df


def _merge_parents(names: list[str]) -> list[list[int]]:
    """Which already-merged columns each merge column depends on (its 'parents').

    Default hierarchy is strictly left-to-right, BUT interval columns are special:
    a time point (Interval 1/2) repeats across every property measured at it, so
    gating it behind Data Column / Unit would keep it from merging on a multi-
    property task. Instead an interval merges on its own repeated values, gated
    only by the OUTERMOST key (Data Template / Property Task) and any interval
    column to its left - so it still respects the top-level block boundary but
    spans across the different properties inside one block."""
    parents: list[list[int]] = []
    nonint: list[int] = []
    inte: list[int] = []
    for i, nm in enumerate(names):
        if str(nm).startswith("Interval "):
            parents.append(([nonint[0]] if nonint else []) + inte[:])
            inte.append(i)
        else:
            parents.append(list(nonint))
            nonint.append(i)
    return parents


def _merge_runs(
    rows: list[list[str]], n_merge: int, parents: list[list[int]] | None = None
) -> list[list[int]]:
    """For the first n_merge columns, compute the rowspan of each cell.
    span[r][c] = number of rows this cell spans (0 = absorbed by the cell above).
    `parents[c]` lists the columns that must still be merged for column c to keep
    a run going; when omitted it is every column to the left (a plain hierarchy),
    so 'Cobb Value' under a new Data Template starts a fresh merge."""
    n = len(rows)
    span = [[1] * n_merge for _ in range(n)]
    if parents is None:
        parents = [list(range(c)) for c in range(n_merge)]
    for c in range(n_merge):
        gate = parents[c]
        r = 0
        while r < n:
            k = r + 1
            while (
                k < n
                and rows[k][c] == rows[r][c]
                and rows[k][c] != ""
                and all(span[k][cc] == 0 for cc in gate)  # required parents still merged
            ):
                k += 1
            span[r][c] = k - r
            for j in range(r + 1, k):
                span[j][c] = 0
            r = k
    return span


def _merged_html(
    disp: pd.DataFrame,
    merge_cols: list[str],
    freeze: int,
    max_height: int = 620,
    key_width: int = 220,
    data_width: int = 150,
) -> str:
    """A real merged-cell table (HTML rowspan), like Excel's Merge & Center.
    Streamlit's grid cannot merge cells, so the merged view is rendered as HTML."""
    cols = list(disp.columns)
    n_merge = len([c for c in merge_cols if c in cols])
    merge_idx = [cols.index(c) for c in merge_cols if c in cols]
    body = disp.astype(str).values.tolist()
    # reorder so the merge columns come first in the span computation
    reord = merge_idx + [i for i in range(len(cols)) if i not in merge_idx]
    rows = [[r[i] for i in reord] for r in body]
    hdr = [cols[i] for i in reord]
    span = _merge_runs(rows, n_merge, _merge_parents(hdr[:n_merge]))

    css = (
        "<style>"
        # table-layout:fixed makes the colgroup widths authoritative, so the width
        # controls can BOTH grow and shrink a column (max-content only ever grew).
        ".mtbl{border-collapse:collapse;font-size:13px;table-layout:fixed}"
        ".mtbl th,.mtbl td{border:1px solid #d9d9d9;padding:5px 9px;vertical-align:middle;"
        "overflow:hidden;word-break:break-word}"
        ".mtbl th{background:#f2f2f2;position:sticky;top:0;z-index:3;text-align:left}"
        ".mtbl td.k{background:#fbfbfb;font-weight:500}"
        ".mtbl .stick{position:sticky;background:#fff;z-index:2}"
        ".mtbl th.stick{z-index:4;background:#f2f2f2}"
        ".mwrap{overflow:auto;max-height:" + str(max_height) + "px;"
        "border:1px solid #e6e6e6;border-radius:6px}"
        "</style>"
    )

    # sticky left offsets for the frozen columns; column widths are user-adjustable
    widths = [key_width if i < n_merge else data_width for i in range(len(hdr))]
    total_w = sum(widths)
    offs, acc = [], 0
    for w in widths:
        offs.append(acc)
        acc += w

    def cls(i):
        return " stick" if i < freeze else ""

    def sty(i):
        return f' style="left:{offs[i]}px;min-width:{widths[i]}px"' if i < freeze else ""

    h = [css, f'<div class="mwrap"><table class="mtbl" style="width:{total_w}px">']
    # <colgroup> sets an explicit width for every column so the width controls take
    # effect even for the non-frozen columns (which get no per-cell min-width).
    h.append("<colgroup>")
    for w in widths:
        h.append(f'<col style="width:{w}px;min-width:{w}px">')
    h.append("</colgroup>")
    h.append("<thead><tr>")
    for i, c in enumerate(hdr):
        h.append(f'<th class="{cls(i).strip()}"{sty(i)}>{c}</th>')
    h.append("</tr></thead><tbody>")
    for r in range(len(rows)):
        h.append("<tr>")
        for i, c in enumerate(hdr):
            if i < n_merge:
                s = span[r][i]
                if s == 0:
                    continue
                rs = f' rowspan="{s}"' if s > 1 else ""
                h.append(f'<td class="k{cls(i)}"{sty(i)}{rs}>{rows[r][i]}</td>')
            else:
                h.append(f'<td class="{cls(i).strip()}"{sty(i)}>{rows[r][i]}</td>')
        h.append("</tr>")
    h.append("</tbody></table></div>")
    return "".join(h)


def show_df(
    df: pd.DataFrame,
    key_labels: list[str],
    table_key: str,
    row_ids: list[str] | None = None,
    merge_cols: list[str] | None = None,
):
    """Render a section table with per-table controls.

    Column labels are SINGLE STRINGS - tuple (code, description) labels became a
    pandas MultiIndex, which Streamlit stringifies for long names, leaking the raw
    tuple repr into cells. The description rides along as a header tooltip and an
    optional first row.
    """
    if df.empty:
        st.info("No rows to display.")
        return

    # MERGE ORDER: outermost hierarchy first. 'Name' is the leaf (unique per row);
    # if it led the list, every run would break immediately and nothing would merge -
    # which is why Group/Subgroup never merged before.
    merge_cols = merge_cols or [k for k in key_labels if k != "Name"]

    row_ids = row_ids or [str(i) for i in range(len(df))]
    sel_key, applied_key = f"sel::{table_key}", f"applied::{table_key}"
    sel: dict[str, bool] = st.session_state.setdefault(sel_key, {})
    for rid in row_ids:
        sel.setdefault(rid, True)

    # ---- controls -----------------------------------------------------------
    # The three row-selection buttons are kept tight together on the left; the
    # view controls (freeze / merge / hide / full screen) follow after a gap.
    b1, b2, b3, _gap, c_fz, c_mg, c_hide, c_full = st.columns(
        [1, 1, 1, 0.4, 1.3, 1.0, 2.0, 0.9]
    )
    with b1:
        if st.button("Select all", key=f"sa::{table_key}"):
            for rid in row_ids:
                sel[rid] = True
            st.session_state[applied_key] = False  # bring every row back into view
            st.rerun()
    with b2:
        if st.button("Unselect all", key=f"ua::{table_key}"):
            for rid in row_ids:
                sel[rid] = False
            st.rerun()
    with b3:
        if st.button("Apply selection", key=f"ap::{table_key}", type="primary"):
            st.session_state[applied_key] = True
            st.rerun()
    with c_fz:
        freeze = st.number_input(
            "Freeze columns",
            min_value=0,
            max_value=len(key_labels) + len(col_tuples),
            value=len(key_labels),
            step=1,
            key=f"fz::{table_key}",
            help="How many columns stay pinned on the left while you scroll sideways.",
        )
    with c_mg:
        merge = st.checkbox(
            "Merge cells",
            value=False,
            key=f"mg::{table_key}",
            help="Combine repeated Group / Subgroup / Interval cells into one spanning "
            "cell (Excel-style). The merged view is read-only - untick to edit the "
            "row selection.",
        )
    # 'Hide columns' and 'Full screen' now apply in BOTH the interactive and the
    # merged views. The dropdown lists EVERY column - the key columns (Group /
    # Subgroup / Name / ...) as well as the experiment columns - so any of them can
    # be hidden.
    hideable = list(key_labels) + [t[0] for t in col_tuples]
    with c_hide:
        hidden_cols = st.multiselect(
            "Hide columns",
            hideable,
            key=f"hide::{table_key}",
            help="Hide any column from this table (key columns included).",
        )
    with c_full:
        full = st.checkbox(
            "Full screen",
            key=f"full::{table_key}",
            help="Expand this table to (almost) the full window height.",
        )
    hidden_set = set(hidden_cols)

    applied = st.session_state.get(applied_key, False)

    # ---- build the display frame --------------------------------------------
    disp = df.copy()
    disp.columns = [c if isinstance(c, str) else str(c[0]) for c in disp.columns]
    # Arrow needs one type per column (Trial is an int, the description row is text)
    disp = disp.astype(str).replace({"None": "", "nan": "", "<NA>": ""})
    disp.insert(0, "✓", [bool(sel.get(rid, True)) for rid in row_ids])
    disp["__rid__"] = row_ids

    if applied:
        disp = disp[disp["✓"]]
        if disp.empty:
            st.warning("No rows selected. Press **Select all** to bring them all back.")
            return

    if show_desc_row and col_tuples:
        head = {c: "" for c in disp.columns}
        head["✓"] = False
        head[key_labels[0]] = "Description"
        head["__rid__"] = "__desc__"
        for code, desc in col_tuples:
            if code in disp.columns:
                head[code] = desc
        disp = pd.concat([pd.DataFrame([head]), disp], ignore_index=True)

    # ---- decimals: round the experiment (value) columns for display ----------
    if DECIMALS is not None:
        for _code, _ in col_tuples:
            if _code in disp.columns:
                disp[_code] = disp[_code].map(lambda x: _apply_decimals_text(x, DECIMALS))

    # ---- MERGED (read-only, real spanning cells) ------------------------------
    if merge:
        body = disp.drop(columns=["✓", "__rid__"])
        shown = [c for c in body.columns if c not in hidden_set]
        body = body[shown]
        # Column widths are adjustable in the merged view.
        w1, w2 = st.columns(2)
        with w1:
            key_w = st.number_input(
                "Key column width (px)",
                min_value=60,
                max_value=600,
                value=220,
                step=10,
                key=f"kw::{table_key}",
                help="Width of the merged key columns (Group / Subgroup / Name / ...).",
            )
        with w2:
            dat_w = st.number_input(
                "Data column width (px)",
                min_value=50,
                max_value=400,
                value=150,
                step=10,
                key=f"dw::{table_key}",
                help="Width of the experiment (value) columns.",
            )
        st.markdown(
            _merged_html(
                body,
                [c for c in merge_cols if c in body.columns],
                int(freeze),
                max_height=880 if full else 620,
                key_width=int(key_w),
                data_width=int(dat_w),
            ),
            unsafe_allow_html=True,
        )
        st.caption(
            "Merged view is read-only. The XLSX export applies the same merges. "
            "Untick **Merge cells** to change the row selection."
        )
        return

    # ---- INTERACTIVE (checkboxes + pinned columns) ---------------------------
    ordered = [c for c in disp.columns if c != "__rid__" and c not in hidden_set]
    cfg: dict[str, Any] = {
        "✓": st.column_config.CheckboxColumn(
            "", help="Tick the rows to keep, then press Apply selection.", pinned=True
        ),
        "__rid__": None,  # hidden
    }
    for i, c in enumerate([x for x in ordered if x != "✓"], start=1):
        desc = next((d for code, d in col_tuples if code == c), None)
        cfg[c] = st.column_config.Column(label=c, help=desc or None, pinned=i <= freeze)

    # Only pass an explicit height for the full-screen view; omitting it lets the
    # grid auto-size (passing height=None raises on some Streamlit versions).
    editor_kwargs: dict[str, Any] = {"height": 800} if full else {}
    edited = st.data_editor(
        disp,
        use_container_width=True,
        hide_index=True,
        column_config=cfg,
        column_order=ordered,
        disabled=[c for c in ordered if c != "✓"],
        key=f"ed::{table_key}",
        **editor_kwargs,
    )

    # persist ticks (ignore the synthetic description row)
    for _, r in edited.iterrows():
        rid = r.get("__rid__")
        if rid and rid != "__desc__":
            sel[rid] = bool(r["✓"])


# ===========================================================================
# 6) Results drill-down (lazy, per selected Property Task) + interval resolve
# ===========================================================================
def _column_value(col) -> str:
    """Recorded value of one PropertyValue. Albert stores it in the NESTED
    PropertyData object (col.property_data.value) - the top-level value/
    valueNumeric/valueString fields are usually empty on a GET."""
    pdat = getattr(col, "property_data", None)
    for src in (pdat, col):
        if src is None:
            continue
        for attr in ("value", "numeric_value", "string_value"):
            v = getattr(src, attr, None)
            if v not in (None, ""):
                return str(v)
    return ""


def _unit_name(col) -> str:
    u = getattr(col, "unit", None)
    if isinstance(u, dict):
        return str(u.get("name") or u.get("Name") or "")
    return str(getattr(u, "name", "") or "")


def _records_from_tpds(
    tpds,
    task_name: str = "",
    task_id: str = "",
    wf_of_block: dict[str, str] | None = None,
    task_workflows: list[str] | None = None,
) -> list[dict]:
    wf_of_block = wf_of_block or {}
    task_workflows = task_workflows or []
    recs: list[dict] = []
    for tpd in tpds:
        dt = getattr(tpd, "data_template", None)
        dt_name = getattr(dt, "name", None) or getattr(dt, "id", "") or "(no template)"
        # PropertyDataInventoryInformation exposes `.inventory_id` (alias "id")
        inv = getattr(tpd, "inventory", None)
        inv_id = getattr(inv, "inventory_id", None) or getattr(inv, "id", None)
        lot_id = getattr(inv, "lot_id", None) or ""
        block_id = getattr(tpd, "block_id", None) or ""
        # block -> workflow is authoritative; the links on the property data are a fallback
        wf = getattr(tpd, "initial_workflow", None) or getattr(tpd, "finial_workflow", None)
        wf_id = wf_of_block.get(block_id) or (getattr(wf, "id", None) or "")
        for interval in getattr(tpd, "data", None) or []:
            if getattr(interval, "void", False):
                continue
            raw_iv = getattr(interval, "interval_combination", "") or ""
            for trial in getattr(interval, "trials", None) or []:
                if getattr(trial, "void", False):
                    continue
                for col in getattr(trial, "data_columns", None) or []:
                    val = _column_value(col)
                    if val == "":
                        continue
                    recs.append(
                        {
                            "task_id": getattr(tpd, "task_id", None) or task_id,
                            "block_id": block_id,
                            "task_name": task_name,
                            "workflow_id": wf_id,
                            "task_workflows": task_workflows,
                            "Data Template": dt_name,
                            "Data Column": getattr(col, "name", "") or "",
                            "Unit": _unit_name(col),
                            "Trial": getattr(trial, "visible_trial_number", None)
                            or getattr(trial, "trial_number", ""),
                            "raw_interval": raw_iv,
                            "inventory_id": inv_id,
                            "lot_id": lot_id,
                            "value": val,
                        }
                    )
    return recs


def _block_workflow_map(_client: Albert, task_id: str) -> tuple[dict[str, str], list[str]]:
    """(block_id -> workflow_id, all workflow ids on the task).

    TaskPropertyData.InitialWorkflow / FinalWorkflow are frequently null on the
    property-data response, so the workflow id has to come from the task itself:
    Task -> Block -> Workflow -> IntervalCombinations. `tasks.get_by_id` returns a
    hydrated PropertyTask (discriminated union), so `.blocks` is populated.
    The flat list is a fallback for when block ids don't line up."""
    out: dict[str, str] = {}
    all_wf: list[str] = []
    try:
        task = _client.tasks.get_by_id(id=task_id)
    except Exception:  # noqa: BLE001
        return out, all_wf
    for b in getattr(task, "blocks", None) or []:
        bid = getattr(b, "id", None)
        wfs = getattr(b, "workflow", None) or []
        if not isinstance(wfs, list):
            wfs = [wfs]
        for wf in wfs:
            wid = getattr(wf, "id", None)
            if wid:
                all_wf.append(str(wid))
                if bid:
                    out.setdefault(str(bid), str(wid))
    return out, list(dict.fromkeys(all_wf))


def _fetch_task_records(_client: Albert, task: dict) -> list[dict]:
    """One task's property data -> flat records (worker thread).
    Errors are RETURNED, not swallowed - a silent [] is indistinguishable
    from 'task genuinely has no data'."""
    try:
        tpds = _client.property_data.get_all_task_properties(
            task_id=task["id"], with_data_only=True
        )
    except Exception as e:  # noqa: BLE001
        return [{"__error__": f"{type(e).__name__}: {e}", "task_id": task["id"]}]
    wf_of_block, all_wf = _block_workflow_map(_client, task["id"])
    return _records_from_tpds(
        tpds,
        task_name=task["name"],
        task_id=task["id"],
        wf_of_block=wf_of_block,
        task_workflows=all_wf,
    )


ROW_TOKEN_RE = re.compile(r"ROW\d+")


def _workflow_raw(_client: Albert, wf_id: str) -> dict | None:
    """Raw workflow JSON, bypassing the SDK's typed model.

    Some workflows in this tenant carry placeholder intervals that have a Unit
    but no `value`. The SDK's Interval model marks `value` as required, so
    `workflows.get_by_id()` raises a pydantic ValidationError on an otherwise
    well-formed response - and with `get_by_ids()` one bad workflow kills the whole
    batch. So we go to the wire ourselves and read the dict."""
    try:
        resp = _client.session.get(f"/api/v3/workflows/{wf_id}")
        data = resp.json()
    except Exception:  # noqa: BLE001
        return None
    if isinstance(data, dict) and "Items" in data:
        items = data.get("Items") or []
        return items[0] if items else None
    return data if isinstance(data, dict) else None


def _workflow_interval_map(_client: Albert, wf_id: str) -> dict[str, list[str]]:
    """token -> ordered axis setpoints, e.g.
       "ROW3XROW22" -> ["Time: 0 day", "Speed: 20 RPM"]

    SOURCE OF TRUTH: `IntervalCombinations[]`, where each entry has
        interval        "ROW3XROW22"                  <- the token on property data
        intervalParams  "INT1XINT1"                   <- per-parameter 1-based index
        intervalString  "Time: 0 day,Speed: 20 RPM"   <- RESOLVED, ORDERED setpoints

    The three fields are positionally aligned, so the Nth ROW token belongs to the
    Nth comma-segment of intervalString. Axis identity is READ, never assumed.

    Two things I previously got wrong, both disproved by the live payload:
      * `Parameters[].Intervals[].rowId` is NOT a reliable source - on WFL446095 the
        interval objects carry no rowId at all, so that map came out empty and every
        token stayed raw.
      * A ROW token is NOT a positional index and NOT the parameter's own
        `prgPrmRowId`: Time's parameter row is ROW1, yet its four interval values
        carry the tokens ROW3-ROW6. Tokens are workflow-scoped ids per interval VALUE.
    """
    out: dict[str, list[str]] = {}

    combos: list = []
    try:  # typed path first
        wf = _client.workflows.get_by_id(id=wf_id)
        for ic in getattr(wf, "interval_combinations", None) or []:
            combos.append(
                {
                    "interval": getattr(ic, "interval_id", None),
                    "intervalString": getattr(ic, "interval_string", None),
                    "intervalDetails": getattr(ic, "interval_details", None),
                }
            )
    except Exception:  # noqa: BLE001  (ValidationError on placeholder intervals)
        raw = _workflow_raw(_client, wf_id)
        if raw:
            combos = raw.get("IntervalCombinations") or []

    for c in combos:
        token = c.get("interval") if isinstance(c, dict) else None
        if not token:
            continue
        token = str(token)
        # X-SAFE: never str.split("X") - a unit or parameter name may contain an X.
        # ROW tokens are always ROW<int>, so a regex findall is safe.
        row_tokens = ROW_TOKEN_RE.findall(token)

        istr = c.get("intervalString") if isinstance(c, dict) else None
        axes: list[str] = []
        if istr:
            # split on "," between axes, then the FIRST ":" between name and value
            for seg in str(istr).split(","):
                name, _, val = seg.partition(":")
                name, val = name.strip(), val.strip()
                axes.append(f"{name}: {val}" if name and val else (val or name))
        else:
            details = c.get("intervalDetails") or []
            for d in details:
                nm = d.get("name") if isinstance(d, dict) else getattr(d, "name", "")
                vl = d.get("value") if isinstance(d, dict) else getattr(d, "value", "")
                axes.append(f"{nm}: {vl}".strip(": "))

        # N-AXIS SAFE: pair positionally; works for 1, 2 or more crossed axes
        if axes:
            out[token] = axes[: len(row_tokens)] if row_tokens else axes
    return out


def _n_axes(records: list[dict]) -> int:
    """How many Interval columns this data actually needs. Never assume 2:
    a block you believe is single-axis can still emit crossed tokens."""
    n = 0
    for r in records:
        if "__error__" in r:
            continue
        n = max(n, len(ROW_TOKEN_RE.findall(str(r.get("raw_interval", "") or ""))))
    return max(1, n)


def resolve_intervals(records: list[dict]) -> None:
    """Attach 'Interval 1'..'Interval N' from the workflow's IntervalCombinations.
    Recomputed on every render, so a stale resolution can never survive."""
    cache: dict[str, dict[str, list[str]]] = st.session_state.setdefault("wf_intervals", {})
    unresolved: dict[str, str] = st.session_state.setdefault("wf_unresolved", {})

    wanted = sorted(
        {
            w
            for r in records
            for w in ([r.get("workflow_id", "")] + list(r.get("task_workflows") or []))
            if w and w not in cache
        }
    )
    if wanted:
        from concurrent.futures import ThreadPoolExecutor

        with ThreadPoolExecutor(max_workers=min(8, len(wanted))) as ex:
            for wid, m in zip(
                wanted, ex.map(lambda w: _workflow_interval_map(client, w), wanted)
            ):
                cache[wid] = m

    n_axes = max(_n_axes(records), N_AXES)
    for r in records:
        raw = str(r.get("raw_interval", "") or "")
        cands = [
            c for c in ([r.get("workflow_id", "")] + list(r.get("task_workflows") or [])) if c
        ]
        axes: list[str] = []
        if raw:
            for wid in cands:
                hit = cache.get(wid, {}).get(raw)
                if hit:
                    axes = list(hit)
                    break
            if not axes:
                unresolved[raw] = f"workflows tried: {cands or '(none on block or task)'}"
                axes = ROW_TOKEN_RE.findall(raw) or [raw]  # show the raw token, flagged
        for i in range(n_axes):
            r[f"Interval {i + 1}"] = axes[i] if i < len(axes) else ""
        r["_n_axes"] = n_axes


def load_selected_results(_client: Albert, pid: str) -> dict[str, list[dict]]:
    """User picks Property Tasks; only those are fetched (README §12 - the
    per-task endpoint is the fastest bulk path). Loaded tasks stay cached in
    session_state."""
    store_key = f"results_store::v3::{pid}"
    store: dict[str, list[dict]] = st.session_state.setdefault(store_key, {})

    tasks = property_tasks
    if not tasks:
        st.warning("No Property Tasks found in this project.")
        return store

    label_of = {f"{t['name']}  [{t['id']}]": t for t in tasks}
    selected = st.multiselect(
        f"Select the Property Tasks to load ({len(tasks)} available)",
        list(label_of.keys()),
        help="Only the selected tasks are downloaded - one API call each.",
    )
    to_fetch = [label_of[l] for l in selected if label_of[l]["id"] not in store]

    b1, b2, b3 = st.columns([1.4, 1.4, 1])
    with b1:
        if to_fetch and st.button(f"⬇️ Load {len(to_fetch)} selected task(s)", type="primary"):
            _load_tasks(_client, store, to_fetch)
    with b2:
        stale = [label_of[l] for l in selected if label_of[l]["id"] in store]
        if stale and st.button("🔄 Reload selected (discard cache)"):
            for t in stale:
                store.pop(t["id"], None)
            st.session_state["wf_intervals"] = {}
            st.session_state["wf_unresolved"] = {}
            _load_tasks(_client, store, stale)
    with b3:
        st.session_state["fetch_workers"] = st.number_input(
            "Parallel requests",
            min_value=1,
            max_value=48,
            value=int(st.session_state.get("fetch_workers", 16)),
            step=4,
            help="Albert makes one request per (block x formulation). Raising this "
            "shortens the load roughly proportionally. Back off if you see errors.",
        )
    if st.session_state.get("fetch_errors"):
        st.warning(f"Some requests failed: {st.session_state['fetch_errors'][0]}")

    return {
        label_of[l]["id"]: store[label_of[l]["id"]]
        for l in selected
        if label_of[l]["id"] in store
    }


def _load_tasks(_client: Albert, store: dict, tasks_to_fetch: list[dict]) -> None:
    """SPEED FIX. `property_data.get_all_task_properties(task_id)` is a SERIAL
    N+1 inside the SDK: it calls check_for_task_data once, then fires ONE HTTP GET
    per (block x inventory) combination, one after another. For a task with 11
    blocks x 95 formulations that is ~1,000 sequential round-trips - which is why
    loading a Property Block took minutes.

    We keep the exact same endpoints (so units, lots, block ids and trials are all
    preserved) but drive them ourselves: enumerate the combos for every selected
    task, then fan out ALL of them across one shared thread pool. Wall time drops
    roughly by the worker count.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    workers = int(st.session_state.get("fetch_workers", 16))
    prog = st.progress(0.0, text="Enumerating blocks...")

    # --- phase 1: per task, which (block, inventory, lot) combos have data? ----
    def _combos(task: dict):
        try:
            checks = _client.property_data.check_for_task_data(task_id=task["id"])
        except Exception as e:  # noqa: BLE001
            return task, None, [], f"{type(e).__name__}: {e}"
        wf_of_block, all_wf = _block_workflow_map(_client, task["id"])
        combos = [c for c in checks if getattr(c, "data_exists", False)]
        return task, (wf_of_block, all_wf), combos, None

    meta: dict[str, tuple] = {}
    jobs: list[tuple[dict, Any]] = []
    with ThreadPoolExecutor(max_workers=min(8, max(1, len(tasks_to_fetch)))) as ex:
        for task, wfinfo, combos, err in ex.map(_combos, tasks_to_fetch):
            if err:
                store[task["id"]] = [{"__error__": err, "task_id": task["id"]}]
                continue
            meta[task["id"]] = wfinfo
            store[task["id"]] = []
            jobs.extend((task, c) for c in combos)

    if not jobs:
        prog.empty()
        st.rerun()

    # --- phase 2: fan every combo out across ONE pool -------------------------
    def _one(job):
        task, c = job
        try:
            tpd = _client.property_data.get_task_block_properties(
                inventory_id=c.inventory_id,
                task_id=task["id"],
                block_id=c.block_id,
                lot_id=getattr(c, "lot_id", None),
            )
        except Exception as e:  # noqa: BLE001
            return task, None, f"{type(e).__name__}: {e}"
        return task, tpd, None

    done, errs = 0, []
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = [ex.submit(_one, j) for j in jobs]
        for fut in as_completed(futures):
            task, tpd, err = fut.result()
            done += 1
            if err:
                errs.append(err)
            elif tpd is not None:
                wf_of_block, all_wf = meta.get(task["id"], ({}, []))
                store[task["id"]].extend(
                    _records_from_tpds(
                        [tpd],
                        task_name=task["name"],
                        task_id=task["id"],
                        wf_of_block=wf_of_block,
                        task_workflows=all_wf,
                    )
                )
            prog.progress(
                done / len(jobs), text=f"Fetching data {done}/{len(jobs)} (x{workers} parallel)"
            )
    if errs:
        st.session_state["fetch_errors"] = errs[:5]
    prog.empty()
    st.rerun()


def _loaded_records() -> list[dict]:
    return [
        r
        for recs in st.session_state.get(f"results_store::v3::{project_id}", {}).values()
        for r in recs
        if "__error__" not in r
    ]


# Interval columns are generated from the DATA, not hard-coded to two. A block you
# believe is single-axis can still emit crossed tokens (ROW3XROW6), and Albert does
# not guarantee only two axes - so the column count follows the tokens.
N_AXES = max(2, _n_axes(_loaded_records()))
INTERVAL_KEYS = [f"Interval {i + 1}" for i in range(N_AXES)]
RESULT_KEYS = ["Data Template", "Data Column", "Unit"] + INTERVAL_KEYS + ["Trial"]
# Keys for the "Merge Results by DT" view: one row per Data Template / Data Column
# / Unit / Interval, pooled across every selected Property Block (Trial dropped, so
# repeated trials collapse into one cell via the aggregation choice).
MERGE_DT_KEYS = ["Data Template", "Data Column", "Unit"] + INTERVAL_KEYS


def results_long_df(records: list[dict]) -> pd.DataFrame:
    """Tidy/long table - one row per recorded value (analysis-ready).
    The Data Templates filter is applied HERE (rows), never to the columns."""
    recs = [r for r in records if "__error__" not in r]
    if flt_result_dts:
        recs = [r for r in recs if r.get("Data Template") in flt_result_dts]
    if not recs:
        return pd.DataFrame()
    resolve_intervals(recs)
    df = pd.DataFrame(recs)
    df["Experiment"] = df["inventory_id"].map(
        lambda i: (invid_to_tuple.get(i) or ("", ""))[0] or _strip_inv(str(i or ""))
    )
    df["Experiment name"] = df["inventory_id"].map(
        lambda i: (invid_to_tuple.get(i) or ("", ""))[1]
    )
    df["Visible (passes filters)"] = df["inventory_id"].isin(invid_to_tuple)

    # Results-only row filters (Data Column comparison + Interval selection).
    dc_f = st.session_state.get("res_dc_filter")
    if dc_f and dc_f[0] != "All":
        df = df[df["Data Column"].map(lambda x: _cmp_pass(x, *dc_f))]
    iv_sel = st.session_state.get("res_interval_filter") or []
    if iv_sel:
        icols = [c for c in df.columns if str(c).startswith("Interval ")]
        if icols:
            df = df[df[icols].apply(lambda row: any(str(v) in iv_sel for v in row), axis=1)]
    return df


def _agg_cell(values, mode: str) -> str:
    """Combine several measurements that land in one (property x experiment) cell.

    mode 'avg' -> numeric mean of the measurements (e.g. 6.12, 6.65, 5.71 -> 6.16),
                  formatted to the same number of decimals as the inputs; falls back
                  to listing when the values are not all numeric.
    otherwise  -> the distinct values joined with ' | ' (original behaviour)."""
    vals = [str(x) for x in values if str(x) != ""]
    if not vals:
        return ""
    if mode == "avg":
        nums = []
        for x in vals:
            try:
                nums.append(float(x.replace(",", ".")))
            except ValueError:
                nums = []
                break  # non-numeric column -> list instead of averaging
        if nums:
            m = sum(nums) / len(nums)
            # keep as many decimals as the inputs carry (handles 6,12 and 6.12)
            decs = [len(x.replace(",", ".").split(".", 1)[1]) for x in vals if ("." in x or "," in x)]
            dec = min(max(decs, default=0), 6)
            return f"{m:.{dec}f}"
    return " | ".join(dict.fromkeys(vals))


def results_drilldown_df(
    records: list[dict],
    include_foreign: bool = False,
    group_keys: list[str] | None = None,
) -> pd.DataFrame:
    """Pivot: DT | DC | Unit | I1 | I2 | Trial rows x visible experiment cols.
    `include_foreign` also shows inventory items filtered out or belonging to
    other sheets (dropped silently before = looked like 'no data').
    `group_keys` overrides the row key (e.g. MERGE_DT_KEYS drops Trial so several
    Property Blocks pool into one row per Data Template / Column / Interval)."""
    keys = group_keys or RESULT_KEYS
    long = results_long_df(records)
    if long.empty:
        return pd.DataFrame()
    agg_mode = st.session_state.get("results_agg_mode", "list")

    tuple_of = dict(invid_to_tuple)
    extra_cols: list[tuple[str, str]] = []
    if include_foreign:
        for inv in long.loc[~long["Visible (passes filters)"], "inventory_id"].dropna().unique():
            t = (_strip_inv(str(inv)), "(filtered out / other sheet)")
            if t not in extra_cols:
                extra_cols.append(t)
            tuple_of[inv] = t

    g = (
        long.groupby(keys + ["inventory_id"], dropna=False, sort=False)["value"]
        .apply(lambda v: _agg_cell(v, agg_mode))
        .reset_index()
    )
    recs = []
    for kv, chunk in g.groupby(keys, dropna=False, sort=False):
        rec = dict(zip(keys, kv if isinstance(kv, tuple) else (kv,)))
        has = False
        for _, r in chunk.iterrows():
            t = tuple_of.get(r["inventory_id"])
            if t:
                rec[t] = r["value"]
                has = True
        if has:
            recs.append(rec)
    if not recs:
        return pd.DataFrame()
    # Keep rows of the same Data Template together (the source order can interleave
    # them, e.g. a Coating Weight row between two Cobb Value rows), so the merged
    # Data Template cell spans them. First-appearance order and the order within a
    # template are both preserved (Python's sort is stable).
    dt_order: dict[str, int] = {}
    for rec in recs:
        dt_order.setdefault(str(rec.get("Data Template", "")), len(dt_order))
    recs.sort(key=lambda rec: dt_order[str(rec.get("Data Template", ""))])
    return pd.DataFrame(recs).reindex(columns=keys + col_tuples + extra_cols).fillna("")


# ===========================================================================
# 7) Render sections (all obey the same visible_cols)
# ===========================================================================
for s in sections:
    st.subheader(s["label"])

    if s["attr"] != "result_design":
        # --- hierarchy provenance: never let an inferred tree pass as fact ---
        if s["max_depth"] == 0:
            st.warning(
                "⚠️ **Group / Subgroup hierarchy unavailable for this section.** "
                f"`GET /api/v3/worksheet/design/.../rows/sequence` returned no tree"
                + (f" ({s['hierarchy_error']})" if s["hierarchy_error"] else "")
                + ". The flat grid response carries no parent, child, depth or "
                "indent field, so depth cannot be recovered from it - a BLK row "
                "followed by another BLK row is ambiguous between *child* and "
                "*sibling*. Rather than guess, the Group columns are left out. "
                "See the Diagnostics panel for the raw payload."
            )
        elif "ONE LEVEL" in s["hierarchy_source"]:
            st.info(
                "ℹ️ Only a single Group level is available for this section "
                "(the sequence endpoint returned no nested subgroups)."
            )

        # --- per-level row filters (Group, Subgroup 1, ...) ------------------
        row_filter: dict[int, list[str]] = {}
        hcols = hier_cols_for(s)
        if hcols:
            fcols = st.columns(len(hcols))
            for lv, (hc, fc) in enumerate(zip(hcols, fcols)):
                opts = sorted(
                    {r["path"][lv] for r in s["rows"] if len(r["path"]) > lv and r["path"][lv]}
                )
                # rows with nothing at this level (top-level headers, shallow
                # branches) - selectable via (None) instead of vanishing
                if any(len(r["path"]) <= lv for r in s["rows"]):
                    opts = [NONE_LABEL] + opts
                with fc:
                    row_filter[lv] = st.multiselect(
                        hc,
                        opts,
                        key=f"rowfilter::{s['attr']}::{lv}",
                        help=f"'{NONE_LABEL}' = rows with no {hc.lower()}.",
                    )

        # --- Product Design: Name filter (>, <, range or contains) -----------
        name_cmp = None
        if s["attr"] == "product_design":
            name_cmp = _cmp_filter_widget(
                "Name filter",
                key=f"namefilter::{s['attr']}",
                help_txt="Filter Product Design rows by their Name "
                "(>, <, Between range, or Contains).",
            )

        sdf, srids = rows_dataframe(s, row_filter, with_ids=True, name_cmp=name_cmp)
        show_df(sdf, key_cols_for(s), table_key=f"sec::{s['attr']}", row_ids=srids)
        continue

    # ----- Results: pick tasks, load only those, drill-down per task --------
    st.caption(
        "Worksheet Property Blocks in this sheet: "
        + " · ".join(f"📦 {r['name']}" for r in s["rows"] if r["name"])
    )
    r1, r2 = st.columns(2)
    with r1:
        include_foreign = st.checkbox(
            "Include experiments filtered out / from other sheets",
            value=False,
            help="A Property Task can hold data for experiments hidden by the "
            "filters above or living on another sheet.",
        )
    with r2:
        long_view = st.checkbox("Long (tidy) view instead of pivot", value=False)

    agg_choice = st.radio(
        "Repeated measurements per property",
        ["List all values (6.12 | 6.65 | 5.71)", "Average"],
        horizontal=True,
        key=f"agg::{s['attr']}",
        help="When one property has several measurements for the same experiment, "
        "either list every value or show their numeric average. Applies to the pivot "
        "view on screen and to the XLSX / CSV (pivot) downloads.",
    )
    st.session_state["results_agg_mode"] = "avg" if agg_choice == "Average" else "list"

    merge_by_dt = st.checkbox(
        "Merge Results by DT",
        value=False,
        key="results_merge_by_dt",
        help="Pool every selected Property Block into ONE table, with one row per "
        "Data Template + Data Column + Interval (regardless of which block it came "
        "from). The XLSX export shows the same single table, with the experiment "
        "columns aligned to the Product Design columns.",
    )

    # --- Results-only filters: Data Column (>, <, range) + Interval ----------
    _recs_for_opts = _loaded_records()
    if _recs_for_opts:
        resolve_intervals(_recs_for_opts)
    iv_options = sorted(
        {
            str(r.get(k))
            for r in _recs_for_opts
            for k in list(r.keys())
            if str(k).startswith("Interval ") and str(r.get(k)).strip()
        }
    )
    st.markdown("**Results filters**")
    st.session_state["res_dc_filter"] = _cmp_filter_widget(
        "Data Column filter",
        key="res_dc",
        help_txt="Filter Results rows by their Data Column (>, <, Between range, or Contains).",
    )
    st.session_state["res_interval_filter"] = st.multiselect(
        "Interval",
        iv_options,
        key="res_interval",
        help="Show only rows recorded at the selected interval(s). Options are the "
        "intervals present in the loaded Property Tasks."
        if iv_options
        else "Load a Property Task below; its intervals will appear here.",
    )

    loaded = load_selected_results(client, project_id)
    task_names = {t["id"]: t["name"] for t in property_tasks}

    if merge_by_dt:
        all_recs: list[dict] = []
        for recs in loaded.values():
            all_recs += [r for r in recs if "__error__" not in r]
        mdf = results_drilldown_df(
            all_recs, include_foreign=include_foreign, group_keys=MERGE_DT_KEYS
        )
        st.caption(
            f"Merged view · {len(loaded)} Property Block(s) pooled into one table by "
            "Data Template / Data Column / Interval."
        )
        if mdf.empty:
            st.info(
                "No results to merge yet. Load at least one Property Task, and check "
                "the Data Template filter / 'Include experiments filtered out'."
            )
        else:
            rids = [
                "|".join(str(mdf.iloc[i][k]) for k in MERGE_DT_KEYS) for i in range(len(mdf))
            ]
            show_df(mdf, MERGE_DT_KEYS, table_key="res::merged_by_dt", row_ids=rids)
        continue

    for task_id, recs in loaded.items():
        task_name = task_names.get(task_id, task_id)
        errors = [r["__error__"] for r in recs if "__error__" in r]
        with st.expander(f"📦 {task_name}  ·  {task_id}", expanded=True):
            if errors:
                st.error(f"API call failed: {errors[0]}")
                continue
            if not recs:
                st.warning(
                    "Albert returned no property data for this task "
                    "(`check_for_task_data` found no combination with data). "
                    "Use the raw-payload expander at the bottom to inspect it."
                )
                continue
            df = results_long_df(recs) if long_view else results_drilldown_df(
                recs, include_foreign=include_foreign
            )
            if df.empty:
                if flt_result_dts:
                    st.write(
                        "No rows in this task match the selected Data Template(s). "
                        "The experiment columns are unaffected."
                    )
                else:
                    st.write(
                        "Data exists, but none of it belongs to the currently visible "
                        "experiments - tick 'Include experiments filtered out'."
                    )
            elif long_view:
                st.dataframe(df, use_container_width=True, hide_index=True)
            else:
                rids = ["|".join(str(df.iloc[i][k]) for k in RESULT_KEYS) for i in range(len(df))]
                show_df(df, RESULT_KEYS, table_key=f"res::{task_id}", row_ids=rids)


# ===========================================================================
# 8) Downloads (respect the global filters)
# ===========================================================================
st.header("4️⃣ Download")


def all_results_df() -> pd.DataFrame:
    store = st.session_state.get(f"results_store::v3::{project_id}", {})
    frames = []
    for task_id, recs in store.items():
        clean = [r for r in recs if "__error__" not in r]
        df = results_drilldown_df(clean, include_foreign=True)
        if not df.empty:
            df.insert(0, "Property Task", clean[0].get("task_name") or task_id)
            frames.append(df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def merged_results_by_dt_df() -> pd.DataFrame:
    """'Merge Results by DT' pooled across every loaded Property Block: one row per
    Data Template / Data Column / Interval, experiment columns shared."""
    store = st.session_state.get(f"results_store::v3::{project_id}", {})
    all_recs: list[dict] = []
    for recs in store.values():
        all_recs += [r for r in recs if "__error__" not in r]
    return results_drilldown_df(all_recs, include_foreign=True, group_keys=MERGE_DT_KEYS)


def all_results_long_df() -> pd.DataFrame:
    store = st.session_state.get(f"results_store::v3::{project_id}", {})
    frames = [results_long_df(recs) for recs in store.values()]
    frames = [f for f in frames if not f.empty]
    if not frames:
        return pd.DataFrame()
    out = pd.concat(frames, ignore_index=True)
    cols = [
        "task_id", "task_name", "block_id", "workflow_id",
        "Data Template", "Data Column", "Unit",
        *INTERVAL_KEYS, "raw_interval", "Trial",
        "Experiment", "Experiment name", "inventory_id", "lot_id",
        "Visible (passes filters)", "value",
    ]
    return out.reindex(columns=[c for c in cols if c in out.columns])


def build_xlsx() -> bytes:
    """Report-ready workbook.

    THE FIX: every section previously wrote its data columns at its own offset
    (Product started after its key columns, Results after 'Property Task' + 6 more),
    so the experiment columns did not line up down the page. Now there is ONE fixed
    grid: a key block of KEY_W columns on the left, then the experiment columns at
    the SAME absolute position for every section. Read straight down column F and
    you are reading one experiment across Product, Process, Results and Apps.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # --- one key block wide enough for the widest section ---------------------
    merge_by_dt = bool(st.session_state.get("results_merge_by_dt"))
    per_section_keys = {s["attr"]: key_cols_for(s) for s in sections}
    per_section_keys["result_design"] = (
        MERGE_DT_KEYS if merge_by_dt else ["Property Task"] + RESULT_KEYS
    )
    KEY_W = max(len(v) for v in per_section_keys.values())
    FIRST_EXP = KEY_W + 1  # 1-based column of the first experiment

    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"

    NAVY = "1F3864"
    GREY = "F2F2F2"
    BAND = "DDEBF7"
    bold_w = Font(bold=True, color="FFFFFF", size=11)
    bold = Font(bold=True)
    ital = Font(italic=True, size=9, color="555555")
    sect_fill = PatternFill("solid", fgColor=NAVY)
    head_fill = PatternFill("solid", fgColor=GREY)
    band_fill = PatternFill("solid", fgColor=BAND)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- title block ----------------------------------------------------------
    ws.cell(row=1, column=1, value=f"Albert Worksheet - {sel_proj}").font = Font(
        bold=True, size=14
    )
    ws.cell(
        row=2,
        column=1,
        value=f"Sheet: {sel_sheet_name}   |   {len(visible_cols)} of {len(exp_cols_all)} "
        f"experiments shown   |   exported {pd.Timestamp.now():%Y-%m-%d %H:%M}",
    ).font = ital

    # --- frozen experiment header (ID over description) -----------------------
    HDR = 4
    for j, (code, desc) in enumerate(col_tuples):
        c1 = ws.cell(row=HDR, column=FIRST_EXP + j, value=code)
        c1.font, c1.alignment, c1.fill, c1.border = bold, ctr, band_fill, box
        c2 = ws.cell(row=HDR + 1, column=FIRST_EXP + j, value=desc)
        c2.font, c2.alignment, c2.border = ital, ctr, box
    ws.cell(row=HDR, column=1, value="Experiment →").font = bold
    ws.freeze_panes = ws.cell(row=HDR + 2, column=FIRST_EXP)

    r = HDR + 2

    def write_section(label: str, keys: list[str], rows_iter, merge_cols: list[str]) -> None:
        nonlocal r
        r += 1
        # full-width section banner
        ws.cell(row=r, column=1, value=label.upper()).font = bold_w
        for cc in range(1, FIRST_EXP + len(col_tuples)):
            ws.cell(row=r, column=cc).fill = sect_fill
        r += 1
        for i, k in enumerate(keys):
            c = ws.cell(row=r, column=1 + i, value=k)
            c.font, c.fill, c.border, c.alignment = bold, head_fill, box, lft
        for j in range(len(col_tuples)):
            c = ws.cell(row=r, column=FIRST_EXP + j)
            c.fill, c.border = head_fill, box
        r += 1

        first_data = r
        keymat: list[list[str]] = []
        for keyvals, expvals in rows_iter:
            padded = [str(keyvals[i]) if i < len(keyvals) else "" for i in range(KEY_W)]
            keymat.append(padded)
            for i in range(KEY_W):
                c = ws.cell(row=r, column=1 + i, value=padded[i])
                c.border, c.alignment = box, lft
            for j, v in enumerate(expvals):
                c = ws.cell(row=r, column=FIRST_EXP + j, value=_num(v))
                c.border, c.alignment = box, ctr
            r += 1

        # --- real Excel merges on the key columns (same runs as the UI) --------
        idxs = [keys.index(m) for m in merge_cols if m in keys]
        if keymat and idxs:
            ordered = [[row[i] for i in idxs] for row in keymat]
            spans = _merge_runs(ordered, len(idxs), _merge_parents([keys[i] for i in idxs]))
            for rr in range(len(keymat)):
                for cc, col_i in enumerate(idxs):
                    s = spans[rr][cc]
                    if s > 1:
                        ws.merge_cells(
                            start_row=first_data + rr,
                            start_column=1 + col_i,
                            end_row=first_data + rr + s - 1,
                            end_column=1 + col_i,
                        )
                        mc = ws.cell(row=first_data + rr, column=1 + col_i)
                        mc.alignment = Alignment(
                            horizontal="left", vertical="center", wrap_text=True
                        )

    def _num(v):
        """Write numbers as numbers so Excel can chart/aggregate them."""
        if isinstance(v, str):
            t = v.strip().replace(",", ".")
            try:
                return float(t) if t not in ("", "-") else v
            except ValueError:
                return v
        return v

    def _apply_row_selection(df: pd.DataFrame, rids: list[str], table_key: str) -> pd.DataFrame:
        """Export exactly what's on screen: if the user pressed Apply selection,
        only the ticked rows go into the workbook."""
        if not st.session_state.get(f"applied::{table_key}", False):
            return df
        sel = st.session_state.get(f"sel::{table_key}", {})
        keep = [i for i, rid in enumerate(rids) if sel.get(rid, True)]
        return df.iloc[keep]

    for s in sections:
        if s["attr"] != "result_design":
            keys = per_section_keys[s["attr"]]
            df, rids = rows_dataframe(s, with_ids=True)
            df = _apply_row_selection(df, rids, f"sec::{s['attr']}")
            write_section(
                s["label"],
                keys,
                (
                    ([row[k] for k in keys], [row[t] for t in col_tuples])
                    for _, row in df.iterrows()
                ),
                merge_cols=[k for k in keys if k != "Name"],  # hierarchy, not the leaf
            )
        else:
            # 'Merge Results by DT': one pooled table (no Property Task column);
            # otherwise the per-task table with Property Task as the outermost key.
            rdf = merged_results_by_dt_df() if merge_by_dt else all_results_df()
            keys = per_section_keys["result_design"]
            write_section(
                s["label"],
                keys,
                (
                    (
                        [row.get(k, "") for k in keys],
                        [row.get(t, "") for t in col_tuples],
                    )
                    for _, row in rdf.iterrows()
                )
                if not rdf.empty
                else iter(()),
                # Merge on all key columns. Property Task (when present) is the
                # outermost, so its cells span all the rows of one task, Excel-style.
                merge_cols=keys,
            )

    # --- widths ---------------------------------------------------------------
    ws.column_dimensions["A"].width = 34
    for i in range(2, KEY_W + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    for j in range(len(col_tuples)):
        ws.column_dimensions[get_column_letter(FIRST_EXP + j)].width = 16
    ws.row_dimensions[HDR + 1].height = 42

    # --- tidy long results on a second sheet (analysis-ready) -----------------
    ldf = all_results_long_df()
    if not ldf.empty:
        ws2 = wb.create_sheet("Results (long)")
        ws2.append(list(ldf.columns))
        for c in ws2[1]:
            c.font, c.fill, c.border = bold, head_fill, box
        for _, row in ldf.iterrows():
            ws2.append([_num(v) for v in row.tolist()])
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions
        for i, cname in enumerate(ldf.columns, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = max(
                12, min(38, len(str(cname)) + 4)
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


d1, d2, d3 = st.columns(3)
with d1:
    _xlsx = build_xlsx()
    if os.environ.get("ALBERT_DUMP_XLSX"):  # offline inspection / tests
        with open(os.environ["ALBERT_DUMP_XLSX"], "wb") as _f:
            _f.write(_xlsx)
    st.download_button(
        "📥 XLSX (filtered worksheet + results)",
        data=_xlsx,
        file_name=f"albert_{project_id}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with d2:
    rdf = all_results_df()
    st.download_button(
        "📥 CSV (results pivot)",
        data=rdf.to_csv(index=False) if not rdf.empty else "",
        file_name=f"albert_{project_id}_results.csv",
        mime="text/csv",
        use_container_width=True,
        disabled=rdf.empty,
    )
with d3:
    ldf = all_results_long_df()
    st.download_button(
        "📥 CSV (results tidy/long)",
        data=ldf.to_csv(index=False) if not ldf.empty else "",
        file_name=f"albert_{project_id}_results_long.csv",
        mime="text/csv",
        use_container_width=True,
        disabled=ldf.empty,
        help="One row per value, with resolved intervals - ready for pandas/PSD analysis.",
    )


# ===========================================================================
# Diagnostics
# ===========================================================================
with st.expander("🔧 Row hierarchy - raw `rows/sequence` payload & resolved paths"):
    for s in sections:
        st.markdown(f"**{s['label']}**")
        st.write(
            {
                "source": s["hierarchy_source"],
                "max depth (ancestor levels)": s["max_depth"],
                "JSON keys actually found": s["hierarchy_keys"] or "(none)",
                "ancestor rowIds with no name found": s.get("hierarchy_unresolved") or "(none)",
                "error": s["hierarchy_error"] or "(none)",
            }
        )
        if s["hierarchy_raw"] is not None:
            st.json(s["hierarchy_raw"], expanded=False)
        paths_df = pd.DataFrame(
            [
                {
                    "row_id": r["row_id"],
                    "type": r["type_raw"].split(".")[-1],
                    "name": r["name"],
                    "depth": r["depth"],
                    "breadcrumb": " > ".join(r["path"] + [r["name"]]),
                }
                for r in s["rows"]
            ]
        )
        st.dataframe(paths_df, use_container_width=True, hide_index=True)
        st.divider()

with st.expander("🔧 Interval resolution (token → setpoint)"):
    unres = st.session_state.get("wf_unresolved", {})
    if unres:
        st.error(
            "**Tokens that could not be resolved** (shown raw in the tables). "
            "Each entry lists the workflows that were searched:"
        )
        st.write(unres)
    else:
        st.success("Every interval token resolved to a setpoint.")
    st.caption(
        f"Interval columns in use: {N_AXES}. Tokens come from "
        "`Workflow.IntervalCombinations[].interval`; the setpoints and their "
        "left-to-right order come from the matching `intervalString`."
    )
    maps = st.session_state.get("wf_intervals", {})
    if maps:
        st.dataframe(
            pd.DataFrame(
                [
                    {"workflow": w, "token": tok, **{f"Interval {i+1}": a for i, a in enumerate(axes)}}
                    for w, m in maps.items()
                    for tok, axes in m.items()
                ]
            ),
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.write("(no workflows loaded yet)")

with st.expander("🔧 Filter sources (facets, tags, predecessor, data templates)"):
    st.write("**Albert inventory facets** (`inventory.get_all_facets`, project-scoped):")
    st.dataframe(
        pd.DataFrame(
            [
                {"facet parameter": p, "value": n, "count": c}
                for p, vals in facets.items()
                for n, c in vals
            ]
        ),
        use_container_width=True,
        hide_index=True,
    )
    st.write(
        "**Per-formulation filter data.** Tags come from `InventoryItem.tags` - the "
        "name is on `Tag.tag`, not `Tag.name` - with any id-only tag resolved via "
        "`tags.get_by_ids`. Predecessor is read from the Apps **PDC** row, the only "
        "place Albert stores it (it is not a field on InventoryItem)."
    )
    st.dataframe(
        pd.DataFrame(
            [
                {
                    "inventory_id": c["inventory_id"],
                    "column_id (Product Design)": c["column_id"],
                    "tags": ", ".join(inv_meta.get(c["inventory_id"], {}).get("tags", [])),
                    "predecessor": inv_meta.get(c["inventory_id"], {}).get("predecessor", ""),
                    "created_by": inv_meta.get(c["inventory_id"], {}).get("created_by", ""),
                    "data_templates": ", ".join(
                        sorted(dts_of_inv.get(c["inventory_id"], set()))
                    ),
                    "locked": c["locked"],
                }
                for c in exp_cols_all
            ]
        ),
        use_container_width=True,
        hide_index=True,
    )
    st.write("**Data template id → name used in the filter:**", dt_name_of or "(none)")

with st.expander("🔧 Raw property-data payload (one task)"):
    if property_tasks:
        pick = st.selectbox(
            "Task", [f"{t['name']}  [{t['id']}]" for t in property_tasks], key="dbg_task"
        )
        tid = pick.split("[")[-1].rstrip("]")
        if st.button("Show raw response"):
            try:
                checks = client.property_data.check_for_task_data(task_id=tid)
                st.write("**check_for_task_data** (drives which combos are fetched):")
                st.dataframe(
                    pd.DataFrame([c.model_dump() for c in checks]), use_container_width=True
                )
                tpds = client.property_data.get_all_task_properties(
                    task_id=tid, with_data_only=True
                )
                st.write(f"**get_all_task_properties** -> {len(tpds)} block/inventory objects")
                for tpd in tpds[:3]:
                    st.json(tpd.model_dump(by_alias=True, mode="json"), expanded=False)
            except Exception as e:  # noqa: BLE001
                st.exception(e)

with st.expander("🔧 Diagnostics"):
    st.write("**Columns** (is_label_col=True are excluded as duplicates of row names):")
    st.dataframe(pd.DataFrame(columns), use_container_width=True, hide_index=True)
    st.write("**Formulation metadata used by the filters** (tags / predecessor / creator):")
    st.dataframe(
        pd.DataFrame(
            [
                {
                    "inventory_id": k,
                    "tags": ", ".join(v["tags"]),
                    "predecessor": v["predecessor"],
                    "created_by": v["created_by"],
                    "name": v["name"],
                    "alias": v["alias"],
                }
                for k, v in inv_meta.items()
            ]
        ),
        use_container_width=True,
        hide_index=True,
    )
    st.write("**Property Tasks & their Data Templates** (drives the DT filter):")
    st.dataframe(
        pd.DataFrame(
            [
                {
                    "task": t["name"],
                    "id": t["id"],
                    "state": t["state"],
                    "data_templates": ", ".join(t["data_templates"]),
                    "n_inventories": len(t["inventory_ids"]),
                }
                for t in property_tasks
            ]
        ),
        use_container_width=True,
        hide_index=True,
    )
    for s in sections:
        st.write(
            f"**{s['label']}** row types:",
            sorted({r["type_raw"] for r in s["rows"]}),
            " | link_id samples:",
            [r["link_id"] for r in s["rows"][:8]],
        )
